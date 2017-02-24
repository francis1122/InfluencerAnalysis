#!/usr/bin/env python
import math
import sys
import array
import openpyxl
import mysql.connector
import time
import datetime
import MySQLdb
import _List
from datetime import datetime, timedelta
from Tkinter import *
from tkFileDialog  import *

print('starting program')

windowWidth = 800

def findIDofInfluencer(influencer):
    cnx = mysql.connector.connect(user='root', password='',
                                  host='127.0.0.1',
                                  database='MediaTracker')

    cursor = cnx.cursor()
    query_string = 'select id from influencer where name = "%s"' % (influencer)
    # cursor.executemany()
    cursor.execute(query_string)

    result = cursor.fetchone()
    influencerID = result[0]
    #print(name)

    cursor.close()
    cnx.close()

    return influencerID

#get array of usernames given array of followerIds
def getAllInfluencers():
    cnx = mysql.connector.connect(user='root', password='',
                                  host='127.0.0.1',
                                  database='MediaTracker')
    cursor = cnx.cursor()

    cursor.execute("select name FROM influencer")
    result = cursor.fetchall()

    cursor.close()
    cnx.close()

    normalize = []
    for n in result:
        normalize.append(n[0])

  #  print(normalize)
    return normalize

def getFollowerCountForUser():
    cnx = mysql.connector.connect(user='root', password='',
                                  host='127.0.0.1',
                                  database='MediaTracker')


    cursor = cnx.cursor()

    cursor.execute("select name FROM influencer")
    result = cursor.fetchall()

    cursor.close()
    cnx.close()

    # normalize = []
    # for n in result:
    #     normalize.append(n[0])

    #  print(normalize)
    return result


#is user currently following specific influencer
# get users that are currently following influencerID
def getActiveFollowerIdsForNetwork(date = int(time.time())):

    cnx = mysql.connector.connect(user='root', password='',
                                  host='127.0.0.1',
                                  database='MediaTracker')
    #20170221
    #1487743026
    #1486246831
    #1486246831
    cursor = cnx.cursor()
    # insert data into followers
    query_string = """select followerid
        from influencer2follower inner
        join(select
        followerid, MAX(created) as created
        from influencer2follower where UNIX_TIMESTAMP(created) < %s group
        by
        followerid) as t1
        using(followerid, created)
        where followed = 1 """ % (date)

    print query_string
    cursor.execute(query_string)

    result = cursor.fetchall()
    #print(result)

    normalize = []
    for n in result:
        normalize.append(n[0])

    cursor.close()
    cnx.close()
    return normalize

def getActiveFollowerIdsForInfluencer(influencerID, date = int(time.time())):

    cnx = mysql.connector.connect(user='root', password='',
                                  host='127.0.0.1',
                                  database='MediaTracker')
    #20170221
    #1487743026
    #1486246831
    #1461913053
    cursor = cnx.cursor()
    # insert data into followers
    query_string = """select followerid
        from influencer2follower inner
        join(select
        followerid, MAX(created) as created
        from influencer2follower where UNIX_TIMESTAMP(created) < %s and influencerid = %s group
        by
        followerid) as t1
        using(followerid, created)
        where  followed = 1 """ % (date, influencerID)

    print query_string
    cursor.execute(query_string)

    result = cursor.fetchall()
    #print(result)

    normalize = []
    for n in result:
        normalize.append(n[0])

    cursor.close()
    cnx.close()
    return normalize

#get array of followerId given array of usernames
def getFollowerIdForUsernames(usernames):
    if len(usernames) == 0:
        return []
    cnx = mysql.connector.connect(user='root', password='',
                                  host='127.0.0.1',
                                  database='MediaTracker')

    cursor = cnx.cursor()

    format_strings = ','.join(['%s'] * len(usernames))
    cursor.execute("select id FROM followers WHERE account_name IN (%s)" % format_strings,
                   tuple(usernames))

    result = cursor.fetchall()


    normalize = []
    for n in result:
        normalize.append(n[0])

    cursor.close()
    cnx.close()
        #  print(normalize)
    return normalize

#get array of usernames given array of followerIds
def getUsernamesForIds(followerIds):
    cnx = mysql.connector.connect(user='root', password='',
                                  host='127.0.0.1',
                                  database='MediaTracker')
    cursor = cnx.cursor()

    format_strings = ','.join(['%s'] * len(followerIds))
    cursor.execute("select account_name FROM followers WHERE id IN (%s)" % format_strings,
                   tuple(followerIds))
    result = cursor.fetchall()

    cursor.close()
    cnx.close()

    normalize = []
    for n in result:
        normalize.append(n[0])

  #  print(normalize)
    return normalize


#is user currently following specific influencer
def findFollowerStatus(followerID, influencerID):

    cnx = mysql.connector.connect(user='root', password='',
                                  host='127.0.0.1',
                                  database='MediaTracker')


    cursor = cnx.cursor()

    # insert data into followers
    query_string = 'select followed from Influencer2Follower' \
                   ' where followerID = %s and influencerID = %s' \
                   ' order by created desc limit 1 '% (followerID, influencerID)
    cursor.execute(query_string)
   # print('findFollowerStatus')
    result = cursor.fetchone()
    isFollowing = 0
    if result != None:
        if len(result) > 0:
            isFollowing = result[0]

    cursor.close()
    cnx.close()

    if isFollowing == 1:
        return True

    #print('findFollowerStatus')
    return False

def addInfluencer(influencerName):
    success = (False, "tuple usage")
    print("addInfluencer! " + influencerName)
    cnx = mysql.connector.connect(user='root', password='',
                                  host='127.0.0.1',
                                  database='MediaTracker')


    cursor = cnx.cursor()
    # insert data into followers

    query_string = ("""INSERT INTO Influencer
                        (name)
                        VALUES ('%s')""")
    try:
        # cursor.execute()
        cursor.execute(query_string % (influencerName,))
        success = (True,  influencerName + " influencer added successfully")
        # queryResult = cursor.executemany(query_string, [(r,) for r in varlist])
    except mysql.connector.Error as err:
        print("Something went wrong: {}".format(err))
        # print("look at this {}".format(queryResult.fetchall()))
        print ('{} is already a follower'.format(influencerName))
        success = (False, influencerName + " failed to get added, probably a duplicated")

    cnx.commit()
    cursor.close()
    cnx.close()
    return success

def addFollowers(followerUsernames):
    cnx = mysql.connector.connect(user='root', password='',
                                  host='127.0.0.1',
                                  database='MediaTracker')


    cursor = cnx.cursor()
    # insert data into followers

    query_string = ("""INSERT INTO followers
                    (account_name)
                    VALUES ('%s')""")
    for r in followerUsernames:
        try:
            # cursor.execute()
            cursor.execute(query_string % (r,))
            # queryResult = cursor.executemany(query_string, [(r,) for r in varlist])
        except mysql.connector.Error as err:
            print("Something went wrong: {}".format(err))
            # print("look at this {}".format(queryResult.fetchall()))
            print ('{} is already a follower'.format(r))
    cnx.commit()
    cursor.close()
    cnx.close()


def readInfluencerSheet(path):
    print('reading in')

    #loads the spreadsheet entirely
    workbook = openpyxl.load_workbook(path)

    #gets the name of a spreadsheet
    sheetName = workbook.get_sheet_names()[0]

    # gets the spreadsheet data
    ws = workbook.get_sheet_by_name(sheetName)
    print ("parsing")
    rows = []
    # get the names of followers
    for row_index in range(ws.max_row):
        rows.append(ws.cell(row=row_index + 1,column=2))

    userNames = []
    for cell in rows:
        if(cell.value):
            userNames.append(cell.value)
            print(cell.value)
        else:
            print('found a null')

    return userNames



def addUnfollowersToInfluencer(unfollowerIds, influencerId, date):
    cnx = mysql.connector.connect(user='root', password='',
                                  host='127.0.0.1',
                                  database='MediaTracker')


    cursor = cnx.cursor()

    # find list of all followers that were just added
    #query = 'select id from followers'

   # cursor.execute(query)
    data = []
    for unfollowerid in unfollowerIds:
        data.append((unfollowerid, influencerId, date, 1))

    # extra saftey, stops from adding duplicate following rows
    # for row in cursor:
    #     # 1 is mario's ID row[0] is followers ID
    #     isFollower = findFollowerStatus(row[0], 1)
    #     if (not isFollower):
    #         data.append((row[0], influencerID, date, 1))

    print (data)

    query_string = 'INSERT INTO Influencer2Follower (followerID, influencerID, created, unfollowed) VALUES (%s, %s, %s, %s)'
    cursor.executemany(query_string, data)
    cnx.commit()

    cursor.close()
    cnx.close()


def addFollowersToInfluencer(followerIds, influencerId, date):
    cnx = mysql.connector.connect(user='root', password='',
                                  host='127.0.0.1',
                                  database='MediaTracker')

    #cursor = cnx.cursor()
    cursor = cnx.cursor(buffered=True)
    # find list of all followers that were just added
  #  query = 'select id from followers'

 #   cursor.execute(query)
    data = []
    for followerid in followerIds:
        data.append((followerid, influencerId, date, 1))

    # extra saftey, stops from adding duplicate following rows
    # for row in cursor:
    #     # 1 is mario's ID row[0] is followers ID
    #     isFollower = findFollowerStatus(row[0], 1)
    #     if (not isFollower):
    #         data.append((row[0], influencerID, date, 1))



    print (data)

    query_string = 'INSERT INTO Influencer2Follower (followerID, influencerID, created, followed) VALUES (%s, %s, %s, %s)'
    cursor.executemany(query_string, data)

  #  cursor.fetchall()
    cnx.commit()

    cursor.close()
    cnx.close()



class App:
    def __init__(self, master):
        master.minsize(width=windowWidth, height=800)
        master.maxsize(width=windowWidth, height=800)

        self.listOfInfluencers = getAllInfluencers()
        print(self.listOfInfluencers )
        self.createTopFrame(master)
        self.createInputFrame(master)
        self.createNetworkOverviewFrame(master)

        # self.newInfluencerNameString = StringVar()
        # self.newInfluencerNameEntry = Entry(frame, textvariable=self.newInfluencerNameString )
        # self.newInfluencerNameEntry.pack(side = LEFT)
        #
        #
        # self.slogan = Button(frame,
        #                      text="Add Influencer",
        #                      command=self.add_influencer)
        # self.slogan.pack(side=LEFT)


    def createTopFrame(self, master):
        frame = Frame(master, width=windowWidth, height=100)
       # frame.pack(fill=None, expand=False)
        frame.pack()
        self.title = Label(frame, text="Influencer Analysis!")
        self.title.pack(fill=None, side=TOP)

        # desciption and user feedback from last action
        # self.userFeedbackString = StringVar()
        self.descriptionLabelText = "welcome"
        self.descriptionLabel = Label(frame, text=self.descriptionLabelText)
        self.descriptionLabel.pack(fill=None, side=TOP)

    def createInputFrame(self, master):
        frame = Frame(master, width=windowWidth, height=150)
       # frame.pack(fill=None, expand=False)
        frame.pack()

        inputLabel = Label(frame, text="Inputs")
        inputLabel.pack(fill=None, side=TOP)

        innerFrame1 = Frame(frame)
        innerFrame1.pack()

        # create adding influencer UI
        self.newInfluencerNameString = StringVar()
        self.newInfluencerNameEntry = Entry(innerFrame1, textvariable=self.newInfluencerNameString)
        self.newInfluencerNameEntry.pack(side = LEFT)

        self.addInfluencerButton = Button(innerFrame1,
                             text="Add Influencer",
                             command=self.addInfluencer)
        self.addInfluencerButton.pack(side = RIGHT)

        # adding data UI
        innerFrame2 = Frame(frame)
        innerFrame2.pack()
        var = StringVar(frame)
        var.set("" )  # initial value


        option = OptionMenu(innerFrame2, var, *self.listOfInfluencers)
        option.pack(side = LEFT)

        self.uploadFilenameLabel = Label(innerFrame2, text="")
        self.uploadFilenameLabel.pack(side=LEFT)

        self.findFileButton= Button(innerFrame2,
                                       text="find file",
                                       command=self.findFilePressed)
        self.findFileButton.pack(side=LEFT)

        self.uploadDataButton = Button(innerFrame2,
                                          text="Upload Data",
                                          command=self.uploadData)
        self.uploadDataButton.pack(side =RIGHT)

    def createNetworkOverviewFrame(self, master):
        frame = Frame(master, width=windowWidth, height=200)
        #frame.pack(fill=None, expand=False)
        frame.pack()

        queriesLabel= Label(frame, text="Queries")
        queriesLabel.pack(fill=None, side=TOP)

        uniqueUserFrame = Frame(frame, width=windowWidth, height=200)
        uniqueUserFrame.pack()

        uniqueLabel = Label(uniqueUserFrame, text= "Network Unique Followers")
        uniqueLabel.grid(row = 0, column = 0)

        nowLabel = Label(uniqueUserFrame, text="Now")
        nowLabel.grid(row=0, column=1)

        nowNetworkFollowerCount = getActiveFollowerIdsForNetwork(date=self.getTimeStampDaysOld(0))
        #followerCount = str(len(nowNetworkFollowerCount))
        nowDataLabel = Label(uniqueUserFrame, text=str(len(nowNetworkFollowerCount)) )
        nowDataLabel.grid(row=1, column=1)

        oneLabel = Label(uniqueUserFrame, text="One Day")
        oneLabel.grid(row=0, column=2)

        oneNetworkFollowerCount = getActiveFollowerIdsForNetwork(date=self.getTimeStampDaysOld(1))
        oneDataLabel = Label(uniqueUserFrame, text=str(len(oneNetworkFollowerCount)))
        oneDataLabel.grid(row=1, column=2)

        threeLabel = Label(uniqueUserFrame, text="Three Days")
        threeLabel.grid(row=0, column=3)

        threeNetworkFollowerCount = getActiveFollowerIdsForNetwork(date=self.getTimeStampDaysOld(300))
        threeDataLabel = Label(uniqueUserFrame, text=str(len(threeNetworkFollowerCount)))
        threeDataLabel.grid(row=1, column=3)

        # adding data UI
        innerFrameFollowerCount = Frame(frame)
        innerFrameFollowerCount.pack()
        self.getFollowerCountString = StringVar(frame)
        self.getFollowerCountString.set("")  # initial value

        self.getFollowerCountOption = OptionMenu(innerFrameFollowerCount, self.getFollowerCountString, *self.listOfInfluencers)
        self.getFollowerCountOption.pack(side=LEFT)

        self.getFollowerCountButton = Button(innerFrameFollowerCount,
                             text="Get Follower Count",
                             command=self.getFollowerCount)
        self.getFollowerCountButton.pack(side=RIGHT)

    def createInfoWindow(self, info):
        window = Toplevel(root)
        w = Text(window)

        w.insert(INSERT, info)
        w.pack()

    #helper function
    def getTimeStampDaysOld(self, daysOld):
        dayOld = datetime.now() - timedelta(days=daysOld)
        unix_time = dayOld.strftime("%s")
        return unix_time

    #button callbacks
    def addInfluencer(self):
        name = self.newInfluencerNameString.get()
        success = addInfluencer(name)
        print("what " + success[1])
        self.descriptionLabel['text'] = success[1]
        self.newInfluencerNameString.set("")


    def findFilePressed(self):
        print ("findFilePressed")

        filename = askopenfilename()
        print (filename)
        #filename = filedialog.askopenfilename()
        #print(len(directory ))
        self.uploadFilenameLabel['text'] = filename
        print ("doing stuff" + filename)

    def uploadData(self):
        print("uploadData")

    def getFollowerCount(self):
        name = self.getFollowerCountString.get()
        print("get follower counts for  " + name)
        influencerID = findIDofInfluencer(name)
        #day old

        print self.getTimeStampDaysOld(1)
        #3 days old
        #threeDaysOld = int(datetime.now() - timedelta(days=3))
        print self.getTimeStampDaysOld(3)

        arrayOfFollowers = getActiveFollowerIdsForInfluencer(influencerID, date = self.getTimeStampDaysOld(3))
        followerCount = str( len(arrayOfFollowers))
        self.createInfoWindow(name + " has " + followerCount + " current followers")
        #print("count of stuff " + test )


root = Tk()
app = App(root)
root.mainloop()

# w1 = Label(root, image=logo).pack(side="right")
# explanation = """At present, only GIF and PPM/PGM
# formats are supported, but an interface
# exists to allow additional image file
# formats to be added easily."""
# w2 = Label(root,
#            justify=LEFT,
#            padx = 10,
#            text=explanation).pack(side="left")

#w.pack()

root.mainloop()

# influencerID = findIDofInfluencer("mario")
#
# list = getFollowerIdsForInfluencer(influencerID)
# currentusers = getUsernamesForIds(list)
# spreadsheetUsers = readInfluencerSheet("download-1.xlsx")

# #spreadsheetUsers.append('mouseman')
# #currentusers.append('dick face')
# # print spreadsheetUsers
# #print currentusers
# # creates list of new followers
# newfollowers =  set(spreadsheetUsers) - set(currentusers)

# #now = time.time()
# now = datetime.datetime(year=2017, month=2, day=1)
# #now = time.mktime(now.timetuple())
# newFollowerIds = getFollowerIdForUsernames(newfollowers)
# addFollowersToInfluencer(newFollowerIds, influencerID, now)

# #creates list of users that have unfollowed
# leavers =  set(currentusers) - set(spreadsheetUsers)
# leaverIds = getFollowerIdForUsernames(leavers)
# addUnfollowersToInfluencer(leaverIds, influencerID, now)

# print 'fuckin a my man'
# print newfollowers
# print newFollowerIds
# print 'fuck these shits'
# print leavers
# print leaverIds
#addFollowers(spreadsheetUsers )


#filename = sys.argv[1]
#filename2 = sys.argv[2]